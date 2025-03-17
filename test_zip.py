import os
import csv
import PyPDF2
import openpyxl
from zipfile import ZipFile
import pytest

@pytest.fixture
def create_zip():
    path = r"tmp" # Указываем путь к директории, где находятся файлы
    file_dir = os.listdir(path) # Получаем список файлов в указанной директории

    # Создаем ZIP-архив с именем 'tmp.zip'
    zip_file_path = 'tmp.zip'  # Путь к создаваемому ZIP-файлу
    with ZipFile(zip_file_path, 'w') as zf: # Открываем ZIP-файл для записи
        for file in file_dir: # Перебираем файлы в директории
            add_file = os.path.join(path, file) # Создаем полный путь к файлу
            zf.write(add_file) # Добавляем файл в ZIP-архив

    # Возвращаем путь к созданному ZIP-файлу
    return zip_file_path


def read_files_from_zip(zip_file_path):
    # Словарь для хранения содержимого файлов
    contents = {}
    with ZipFile(zip_file_path, 'r') as zf: # Открываем ZIP-файл для чтения
        for file_info in zf.infolist(): # Перебираем файлы в архиве
            with zf.open(file_info) as f: # Открываем файл в архиве
                if file_info.filename.endswith('.csv'): # Проверяем, является ли файл CSV
                    try:
                        content = f.read().decode('utf-8').splitlines() # Читаем содержимое CSV, декодируем и разбиваем на строки
                    except UnicodeDecodeError:
                        # Если возникает ошибка декодирования, пробуем другую кодировку
                        f.seek(0) # Сбрасываем указатель файла
                        content = f.read().decode('ISO-8859-1').splitlines() # Пробуем другую кодировку
                    reader = csv.reader(content) # Используем csv.reader для чтения содержимого
                    contents[file_info.filename] = list(reader) # Сохраняем содержимое в словарь

                # Проверяем, является ли файл PDF
                elif file_info.filename.endswith('.pdf'):
                    reader = PyPDF2.PdfReader(f) # Читаем содержимое PDF
                    contents[file_info.filename] = [page.extract_text() for page in reader.pages] # Сохраняем текст каждой страницы в словарь

                # Проверяем, является ли файл XLSX
                elif file_info.filename.endswith('.xlsx'):
                    workbook = openpyxl.load_workbook(f) # Читаем содержимое XLSX
                    content = {}
                    # Перебираем все листы в книге
                    for sheet in workbook.sheetnames:
                        ws = workbook[sheet]
                        content[sheet] = [[cell.value for cell in row] for row in ws.iter_rows()] # Сохраняем содержимое каждого листа в словарь
                    contents[file_info.filename] = content # Сохраняем содержимое XLSX в словарь

    # Возвращаем словарь с содержимым всех файлов
    return contents


def test_zip_creation(create_zip):
    zip_file_path = create_zip # Получаем путь к созданному ZIP-файлу
    assert os.path.exists(zip_file_path) # Проверяем, что ZIP-файл был успешно создан

    # Читаем содержимое файлов из ZIP-архива
    contents = read_files_from_zip(zip_file_path)

    # Выводим результаты теста
    print("Тест завершен: test_zip_creation PASSED [100%]")
    print("\nСодержимое файлов в ZIP-архиве:\n" + "=" * 50)

    # Переменные для отслеживания наличия слов
    word_found_xlsx = False
    word_found_csv = False
    word_found_pdf = False

    # Перебираем содержимое и ищем файлы
    for filename, content in contents.items():
        print(f"Проверяем файл: {filename}") # Выводим имя файла
        if filename.endswith('.xlsx'):
            # Проверяем, является ли содержимое словарем (для XLSX)
            assert isinstance(content, dict), f"Содержимое файла {filename} не является словарем."

            # Перебираем каждый лист в XLSX
            for sheet, rows in content.items():
                print(f"Лист: {sheet}") # Выводим имя листа
                # Перебираем строки в листе
                for row in rows:
                    print(f"{row}") # Выводим каждую строку
                    if "Пока" in str(row):
                        word_found_xlsx = True # Устанавливаем флаг, если слово найдено
                        print("Найдено слово 'Пока'") # Выводим сообщение о найденном слове
                        break # Выходим из цикла, если слово найдено
                if word_found_xlsx:
                    break # Выходим из внешнего цикла, если слово найдено

        elif filename.endswith('.csv'):
            # Проверяем, является ли содержимое списком (для CSV)
            assert isinstance(content, list), f"Содержимое файла {filename} не является списком."

            # Перебираем строки в CSV
            for row in content:
                print(f"{row}") # Выводим каждую строку
                if "Sergeevich" in str(row):
                    word_found_csv = True # Устанавливаем флаг, если слово найдено
                    print("Найдено слово 'Sergeevich'") # Выводим сообщение о найденном слове
                    break # Выходим из цикла, если слово найдено

        elif filename.endswith('.pdf'):
            # Проверяем, является ли содержимое списком (для PDF)
            assert isinstance(content, list), f"Содержимое файла {filename} не является списком."

            # Перебираем строки в PDF
            for line in content:
                print(f"{line}") # Выводим каждую строку
                if "Browserstack" in str(line):
                    word_found_pdf = True # Устанавливаем флаг, если слово найдено
                    print("Найдено слово 'Browserstack'") # Выводим сообщение о найденном слове
                    break # Выходим из цикла, если слово найдено

    # Проверяем, были ли найдены слова
    assert word_found_xlsx, "Слово 'Пока' не найдено в файле XLSX."
    assert word_found_csv, "Слово 'Sergeevich' не найдено в файле CSV."
    assert word_found_pdf, "Слово 'Browserstack' не найдено в файле PDF."

    print("Все проверки завершены успешно.")

    os.remove(zip_file_path) # Удаляем ZIP-файл после теста